import os
import re
import json
import time
import threading
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Output folder ─────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "JarvisExports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Shared state for upload flow ──────────────────────────────────────────
_pending_conversion = threading.Event()
_uploaded_pdf_path  = [None]   # list so thread can mutate it


# ════════════════════════════════════════════════════════════════════════════
#  ANALYSE PDF — detect what's inside
# ════════════════════════════════════════════════════════════════════════════
def analyse_pdf(pdf_path: str) -> dict:
    info = {"pages": 0, "has_tables": False, "has_text": False,
            "table_count": 0, "text_pages": 0}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            info["pages"] = len(pdf.pages)
            for page in pdf.pages:
                tables = page.extract_tables()
                text   = page.extract_text()
                if tables:
                    info["has_tables"] = True
                    info["table_count"] += len(tables)
                if text and text.strip():
                    info["has_text"] = True
                    info["text_pages"] += 1
    except Exception as e:
        print(f"[PDF] Analyse error: {e}")
    return info


# ════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════
_HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
_ALT_FILL      = PatternFill("solid", start_color="D6E4F0")
_HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_NORMAL_FONT   = Font(name="Arial", size=10)
_BORDER_SIDE   = Side(style="thin", color="B0C4DE")
_CELL_BORDER   = Border(left=_BORDER_SIDE, right=_BORDER_SIDE,
                        top=_BORDER_SIDE,  bottom=_BORDER_SIDE)
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _style_header(cell):
    cell.font      = _HEADER_FONT
    cell.fill      = _HEADER_FILL
    cell.alignment = _CENTER
    cell.border    = _CELL_BORDER


def _style_data(cell, alt_row: bool = False):
    cell.font      = _NORMAL_FONT
    cell.alignment = _LEFT
    cell.border    = _CELL_BORDER
    if alt_row:
        cell.fill = _ALT_FILL


def _autofit_columns(sheet, min_w=10, max_w=50):
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, min_w), max_w)


# ════════════════════════════════════════════════════════════════════════════
#  WRITE TABLE TO SHEET
# ════════════════════════════════════════════════════════════════════════════
def _write_table(sheet, table: list, start_row: int = 1) -> int:
    """Writes a 2D list table to the sheet. Returns next available row."""
    if not table:
        return start_row

    # Clean None cells
    cleaned = []
    for row in table:
        cleaned.append([str(c).strip() if c is not None else "" for c in row])

    # Header row
    header = cleaned[0]
    for col_idx, val in enumerate(header, start=1):
        cell = sheet.cell(row=start_row, column=col_idx, value=val)
        _style_header(cell)
    sheet.row_dimensions[start_row].height = 22

    # Data rows
    for row_offset, row in enumerate(cleaned[1:], start=1):
        excel_row = start_row + row_offset
        alt = row_offset % 2 == 0
        for col_idx, val in enumerate(row, start=1):
            # Try numeric coercion
            try:
                numeric = float(val.replace(",", "")) if val else None
                cell_val = int(numeric) if numeric is not None and numeric == int(numeric) else numeric
            except (ValueError, AttributeError):
                cell_val = val
            cell = sheet.cell(row=excel_row, column=col_idx, value=cell_val if cell_val is not None else val)
            _style_data(cell, alt)
        sheet.row_dimensions[excel_row].height = 18

    # Freeze header
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)
    _autofit_columns(sheet)
    return start_row + len(cleaned) + 2   # +2 gap before next table


# ════════════════════════════════════════════════════════════════════════════
#  WRITE RAW TEXT TO SHEET
# ════════════════════════════════════════════════════════════════════════════
def _write_text(sheet, text: str):
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    for row_idx, line in enumerate(lines, start=1):
        # Try to split into columns by common delimiters
        if "\t" in line:
            cols = line.split("\t")
        elif re.search(r"\s{2,}", line):
            cols = re.split(r"\s{2,}", line)
        else:
            cols = [line]

        alt = row_idx % 2 == 0
        for col_idx, val in enumerate(cols, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=val.strip())
            _style_data(cell, alt)
        sheet.row_dimensions[row_idx].height = 16

    _autofit_columns(sheet)


# ════════════════════════════════════════════════════════════════════════════
#  MAIN CONVERTER
# ════════════════════════════════════════════════════════════════════════════
def convert_pdf_to_excel(pdf_path: str, speak_fn) -> str | None:
    if not os.path.exists(pdf_path):
        speak_fn("I couldn't find the uploaded PDF. Please try again.")
        return None

    pdf_name   = os.path.splitext(os.path.basename(pdf_path))[0]
    timestamp  = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"{pdf_name}_{timestamp}.xlsx")

    speak_fn("Analysing the PDF content. Please wait.")
    info = analyse_pdf(pdf_path)
    print(f"[PDF2Excel] Info: {info}")

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    sheets_created = 0

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_tables = []
            all_texts  = {}

            # ── Pass 1: collect everything ────────────────────────────────
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                text   = page.extract_text()

                if tables:
                    for tbl in tables:
                        if tbl and len(tbl) > 1:
                            all_tables.append((page_num, tbl))

                if text and text.strip():
                    all_texts[page_num] = text

            # ── Pass 2: tables → individual sheets ────────────────────────
            if all_tables:
                # Group tables: if many small tables, merge into one sheet per page
                pages_with_tables = {}
                for page_num, tbl in all_tables:
                    pages_with_tables.setdefault(page_num, []).append(tbl)

                for page_num, tables in pages_with_tables.items():
                    sheet_name = f"Page {page_num}" if len(tables) == 1 else f"Page {page_num} Tables"
                    ws = wb.create_sheet(title=sheet_name[:31])
                    current_row = 1
                    for tbl in tables:
                        current_row = _write_table(ws, tbl, start_row=current_row)
                    sheets_created += 1

            # ── Pass 3: text-only pages → "Text Content" sheet ────────────
            if all_texts:
                # Pages that have text but no table
                table_pages = {pn for pn, _ in all_tables}
                text_only   = {pn: t for pn, t in all_texts.items() if pn not in table_pages}

                if text_only:
                    ws_text = wb.create_sheet(title="Text Content")
                    current_row = 1
                    for page_num in sorted(text_only.keys()):
                        # Page label
                        label_cell = ws_text.cell(row=current_row, column=1,
                                                   value=f"── Page {page_num} ──")
                        label_cell.font = Font(name="Arial", bold=True, size=11, color="1F4E79")
                        current_row += 1

                        lines = [l.strip() for l in text_only[page_num].split("\n") if l.strip()]
                        for line in lines:
                            if "\t" in line:
                                cols = line.split("\t")
                            elif re.search(r"\s{2,}", line):
                                cols = re.split(r"\s{2,}", line)
                            else:
                                cols = [line]

                            for col_idx, val in enumerate(cols, start=1):
                                cell = ws_text.cell(row=current_row, column=col_idx, value=val.strip())
                                _style_data(cell)
                            current_row += 1
                        current_row += 1   # gap between pages

                    _autofit_columns(ws_text)
                    sheets_created += 1

            # ── Fallback: if nothing extracted, dump raw text ─────────────
            if sheets_created == 0 and all_texts:
                ws = wb.create_sheet(title="Content")
                _write_text(ws, "\n".join(all_texts.values()))
                sheets_created += 1

        if sheets_created == 0:
            speak_fn("The PDF appears to be empty or scanned with no readable text. I can't convert it.")
            return None

        wb.save(output_path)
        print(f"[PDF2Excel] Saved: {output_path}")
        return output_path

    except Exception as e:
        print(f"[PDF2Excel] Conversion error: {e}")
        speak_fn("Something went wrong during conversion. Please try again.")
        return None


# ════════════════════════════════════════════════════════════════════════════
#  SIGNAL FROM UI — called by Eel when user uploads a file
# ════════════════════════════════════════════════════════════════════════════
def set_uploaded_pdf(path: str):
    _uploaded_pdf_path[0] = path
    _pending_conversion.set()


# ════════════════════════════════════════════════════════════════════════════
#  MAIN HANDLER — called from command.py
# ════════════════════════════════════════════════════════════════════════════
def handlePDFToExcel(speak_fn):
    speak_fn("Sure! Please upload your PDF using the upload button that just appeared.")

    # Signal UI to show upload panel
    try:
        import eel
        eel.showPDFUpload()
    except Exception as e:
        print(f"[PDF2Excel] Eel signal error: {e}")

    # Wait up to 120 seconds for the user to upload
    _pending_conversion.clear()
    _uploaded_pdf_path[0] = None

    got_file = _pending_conversion.wait(timeout=120)

    if not got_file or not _uploaded_pdf_path[0]:
        speak_fn("No file was uploaded. Please try again when you're ready.")
        return

    pdf_path = _uploaded_pdf_path[0]
    speak_fn("Got it! Converting your PDF to Excel now.")

    output_path = convert_pdf_to_excel(pdf_path, speak_fn)

    if output_path:
        fname = os.path.basename(output_path)
        speak_fn(f"Done! Your Excel file {fname} is saved in Documents, JarvisExports. Opening it now.")
        time.sleep(0.5)
        os.startfile(output_path)
